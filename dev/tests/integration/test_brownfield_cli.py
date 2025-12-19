"""Brownfield CLI integration tests (BF-13).

subprocess로 실제 `python -m cmis_cli brownfield ...`를 실행해
MVP 수직 슬라이스(import→preview→validate→commit)가 동작하는지 검증합니다.

주의:
- CMIS_STORAGE_ROOT를 임시 디렉토리로 지정하여 테스트 간 격리합니다.
"""

from __future__ import annotations

import json
import os
import sqlite3
import subprocess
import tempfile
from pathlib import Path

from openpyxl import Workbook


def _artifact_file_path(storage_root: Path, artifact_id: str) -> Path:
    db_path = storage_root / ".cmis" / "db" / "artifacts.db"
    assert db_path.exists()

    conn = sqlite3.connect(str(db_path))
    try:
        cur = conn.execute("SELECT file_path FROM artifacts WHERE artifact_id = ?", (str(artifact_id),))
        row = cur.fetchone()
    finally:
        conn.close()

    assert row and row[0]
    return Path(str(row[0]))


def _import_run_fields(storage_root: Path, import_run_id: str) -> dict:
    db_path = storage_root / ".cmis" / "db" / "brownfield.db"
    assert db_path.exists()

    conn = sqlite3.connect(str(db_path))
    try:
        cur = conn.execute(
            """
            SELECT artifact_ids_json, preview_report_artifact_id, validation_report_artifact_id
            FROM import_runs
            WHERE import_run_id = ?
            """,
            (str(import_run_id),),
        )
        row = cur.fetchone()
    finally:
        conn.close()

    assert row
    artifact_ids_json, preview_id, validation_id = row
    artifact_ids = json.loads(artifact_ids_json or "[]")
    assert isinstance(artifact_ids, list)

    return {
        "artifact_ids": [str(x) for x in artifact_ids],
        "preview_report_artifact_id": str(preview_id) if preview_id is not None else None,
        "validation_report_artifact_id": str(validation_id) if validation_id is not None else None,
    }


def test_brownfield_cli_import_preview_validate_commit(project_root: Path) -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        storage_root = Path(tmpdir)
        env = os.environ.copy()
        env["CMIS_STORAGE_ROOT"] = tmpdir

        csv_path = Path(tmpdir) / "sample.csv"
        csv_path.write_text("name,age\nAlice,30\nBob,40\n", encoding="utf-8")

        # import
        r1 = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "brownfield",
                "import",
                str(csv_path),
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r1.returncode == 0, r1.stderr
        p1 = json.loads(r1.stdout)
        assert p1.get("ok") is True
        imp_id = str(p1.get("import_run_id"))
        assert imp_id.startswith("IMP-")

        # preview
        r2 = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "brownfield",
                "preview",
                imp_id,
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r2.returncode == 0, r2.stderr
        assert "Alice" not in r2.stdout
        assert "Bob" not in r2.stdout
        p2 = json.loads(r2.stdout)
        assert p2.get("ok") is True
        preview = p2.get("preview") or {}
        assert int(preview.get("row_count") or 0) == 2
        assert preview.get("columns") == ["name", "age"]

        # validate
        r3 = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "brownfield",
                "validate",
                imp_id,
                "--policy-mode",
                "reporting_strict",
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r3.returncode == 0, r3.stderr
        assert "Alice" not in r3.stdout
        p3 = json.loads(r3.stdout)
        assert p3.get("policy_decision") == "pass"

        # commit (1st)
        r4 = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "brownfield",
                "commit",
                imp_id,
                "--policy-mode",
                "reporting_strict",
                "--focal-actor-context-base-id",
                "PRJ-testcli",
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r4.returncode == 0, r4.stderr
        p4 = json.loads(r4.stdout)
        assert p4.get("ok") is True
        bundle_id_1 = str(p4.get("bundle_id"))
        assert bundle_id_1.startswith("CUB-")
        assert p4.get("focal_actor_context_id") == "PRJ-testcli-v1"

        # re-run determinism: same file → same CUB (dedupe), PRJ version increments
        r1b = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "brownfield",
                "import",
                str(csv_path),
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r1b.returncode == 0, r1b.stderr
        p1b = json.loads(r1b.stdout)
        assert p1b.get("ok") is True
        imp_id_2 = str(p1b.get("import_run_id"))
        assert imp_id_2.startswith("IMP-")
        assert imp_id_2 != imp_id

        r3b = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "brownfield",
                "validate",
                imp_id_2,
                "--policy-mode",
                "reporting_strict",
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r3b.returncode == 0, r3b.stderr
        p3b = json.loads(r3b.stdout)
        assert p3b.get("policy_decision") == "pass"

        r4b = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "brownfield",
                "commit",
                imp_id_2,
                "--policy-mode",
                "reporting_strict",
                "--focal-actor-context-base-id",
                "PRJ-testcli",
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r4b.returncode == 0, r4b.stderr
        p4b = json.loads(r4b.stdout)
        assert p4b.get("ok") is True
        assert str(p4b.get("bundle_id")) == bundle_id_1
        assert p4b.get("focal_actor_context_id") == "PRJ-testcli-v2"

        # context verify (v1, v2)
        for prj_id in ["PRJ-testcli-v1", "PRJ-testcli-v2"]:
            r5 = subprocess.run(
                [
                    "python3",
                    "-m",
                    "cmis_cli",
                    "context",
                    "verify",
                    prj_id,
                ],
                cwd=str(project_root),
                env=env,
                capture_output=True,
                text=True,
                timeout=60,
            )
            assert r5.returncode == 0, r5.stderr
            p5 = json.loads(r5.stdout)
            assert p5.get("ok") is True

        # leakage guard (stored artifacts): preview/validation report must not contain raw rows
        f1 = _import_run_fields(storage_root, imp_id)
        f2 = _import_run_fields(storage_root, imp_id_2)
        assert f1["artifact_ids"] and f2["artifact_ids"]
        assert f1["artifact_ids"][0] == f2["artifact_ids"][0]

        for art_id in [f1["preview_report_artifact_id"], f1["validation_report_artifact_id"]]:
            assert art_id
            p = _artifact_file_path(storage_root, art_id)
            text = p.read_text(encoding="utf-8")
            assert "Alice" not in text
            assert "Bob" not in text


def test_brownfield_cli_xlsx_import_validate_commit(project_root: Path) -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        storage_root = Path(tmpdir)
        env = os.environ.copy()
        env["CMIS_STORAGE_ROOT"] = tmpdir

        xlsx_path = Path(tmpdir) / "sample.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "name"
        ws["B1"] = "value"
        ws["A2"] = "Alice"
        ws["B2"] = 10
        ws["A3"] = "Bob"
        ws["B3"] = "=SUM(1,2)"  # no cached value in openpyxl-generated files
        wb.save(str(xlsx_path))

        # import
        r1 = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "brownfield",
                "import",
                str(xlsx_path),
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r1.returncode == 0, r1.stderr
        p1 = json.loads(r1.stdout)
        assert p1.get("ok") is True
        imp_id = str(p1.get("import_run_id"))
        assert imp_id.startswith("IMP-")

        # preview (must not leak raw rows)
        r2 = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "brownfield",
                "preview",
                imp_id,
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r2.returncode == 0, r2.stderr
        assert "Alice" not in r2.stdout
        assert "Bob" not in r2.stdout
        p2 = json.loads(r2.stdout)
        preview = p2.get("preview") or {}
        assert preview.get("format") == "xlsx"
        assert int(preview.get("row_count") or 0) == 2

        # validate (xlsx formula cached value warning -> warn_only)
        r3 = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "brownfield",
                "validate",
                imp_id,
                "--policy-mode",
                "exploration_friendly",
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r3.returncode == 0, r3.stderr
        assert "Alice" not in r3.stdout
        p3 = json.loads(r3.stdout)
        assert p3.get("policy_decision") in {"warn_only", "pass"}

        # commit (allow warn_only under exploration_friendly)
        r4 = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "brownfield",
                "commit",
                imp_id,
                "--policy-mode",
                "exploration_friendly",
                "--focal-actor-context-base-id",
                "PRJ-xlsxcli",
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r4.returncode == 0, r4.stderr
        p4 = json.loads(r4.stdout)
        assert p4.get("ok") is True
        assert str(p4.get("bundle_id")).startswith("CUB-")
        assert p4.get("focal_actor_context_id") == "PRJ-xlsxcli-v1"

        # context verify
        r5 = subprocess.run(
            [
                "python3",
                "-m",
                "cmis_cli",
                "context",
                "verify",
                "PRJ-xlsxcli-v1",
            ],
            cwd=str(project_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=60,
        )
        assert r5.returncode == 0, r5.stderr
        p5 = json.loads(r5.stdout)
        assert p5.get("ok") is True

        # leakage guard (stored artifacts): preview/validation report must not contain raw rows
        f = _import_run_fields(storage_root, imp_id)
        for art_id in [f["preview_report_artifact_id"], f["validation_report_artifact_id"]]:
            assert art_id
            p = _artifact_file_path(storage_root, art_id)
            text = p.read_text(encoding="utf-8")
            assert "Alice" not in text
            assert "Bob" not in text
