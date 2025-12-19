"""XLSX ingest (BF-09b, Level 1).

목표:
- 업로드 파일을 ART로 저장
- ImportRun(IMP) 생성
- XLSX decode 결과로 preview summary를 만들고 ART로 저장
- ImportRun에 preview artifact를 attach하여 status=decoded로 갱신

Level 1 Guardrails(요약):
- 수식(formula) 셀의 cached value가 없을 수 있음(data_only=True에서 None)
  - 해당 케이스는 preview에 count로만 반영합니다(원문/수치 노출 금지).

주의(누출 방지):
- preview summary는 원문 행/셀 값을 포함하지 않고 shape/통계 중심으로 구성합니다.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

import sqlite3

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from cmis_core.brownfield.db import migrate_brownfield_db, open_brownfield_db
from cmis_core.brownfield.import_run_store import ImportRunStore
from cmis_core.brownfield.uow import UnitOfWork
from cmis_core.stores.artifact_store import ArtifactStore


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _is_non_empty_value(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() != ""
    return True


@dataclass(frozen=True)
class XlsxSheetPreview:
    name: str
    row_count: int
    columns: List[str]
    non_empty_counts: Dict[str, int]
    formula_cell_count: int
    formula_missing_cached_value_count: int
    truncated: Dict[str, Any]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": str(self.name),
            "row_count": int(self.row_count),
            "columns": list(self.columns),
            "non_empty_counts": dict(self.non_empty_counts),
            "formula_cell_count": int(self.formula_cell_count),
            "formula_missing_cached_value_count": int(self.formula_missing_cached_value_count),
            "truncated": dict(self.truncated),
        }


@dataclass(frozen=True)
class XlsxPreviewSummary:
    """XLSX preview summary.

    계약:
    - top-level에 columns/row_count/non_empty_counts를 제공하여 validate/commit(MVP) 경로와 호환
    - sheets에 시트별 summary를 포함(누출 없는 shape 정보만)
    """

    format: str
    primary_sheet: Optional[str]
    columns: List[str]
    row_count: int
    non_empty_counts: Dict[str, int]
    formula_cell_count: int
    formula_missing_cached_value_count: int
    sheets: List[XlsxSheetPreview]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "format": str(self.format),
            "primary_sheet": str(self.primary_sheet) if self.primary_sheet is not None else None,
            "columns": list(self.columns),
            "row_count": int(self.row_count),
            "non_empty_counts": dict(self.non_empty_counts),
            "formula_cell_count": int(self.formula_cell_count),
            "formula_missing_cached_value_count": int(self.formula_missing_cached_value_count),
            "sheets": [s.to_dict() for s in self.sheets],
        }


def _decode_xlsx_to_preview(
    path: Path,
    *,
    max_rows: int = 20000,
    max_cols: int = 200,
) -> XlsxPreviewSummary:
    """XLSX 파일을 읽어 preview용 요약만 생성합니다.

    - 원문/값은 저장하지 않습니다.
    - column 이름은 시트의 header row를 추정하지 않고, A/B/C... column letter로 고정합니다.
    """

    p = Path(path)

    wb_formula = load_workbook(filename=str(p), read_only=True, data_only=False)
    wb_value = load_workbook(filename=str(p), read_only=True, data_only=True)

    sheets: List[XlsxSheetPreview] = []

    primary_sheet_name: Optional[str] = None
    primary_columns: List[str] = []
    primary_row_count = 0
    primary_non_empty: Dict[str, int] = {}
    primary_formula_count = 0
    primary_missing_cached = 0

    for ws in wb_formula.worksheets:
        name = str(ws.title)
        if primary_sheet_name is None:
            primary_sheet_name = name

        ws_val = wb_value[name]

        max_row = int(ws.max_row or 0)
        max_col = int(ws.max_column or 0)

        truncated = {
            "rows": False,
            "cols": False,
            "max_rows_limit": int(max_rows),
            "max_cols_limit": int(max_cols),
        }

        if max_row > max_rows:
            max_row = max_rows
            truncated["rows"] = True
        if max_col > max_cols:
            max_col = max_cols
            truncated["cols"] = True

        columns = [get_column_letter(i) for i in range(1, max_col + 1)]
        non_empty_counts = {c: 0 for c in columns}

        row_count = 0
        formula_cell_count = 0
        missing_cached_count = 0

        if max_row >= 2 and max_col >= 1:
            formula_rows = ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=False)
            value_rows = ws_val.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=True)

            for f_row, v_row in zip(formula_rows, value_rows):
                row_has_any = False
                for idx, col in enumerate(columns):
                    v = v_row[idx] if idx < len(v_row) else None
                    if _is_non_empty_value(v):
                        non_empty_counts[col] = int(non_empty_counts.get(col, 0)) + 1
                        row_has_any = True

                    # formula cached value guardrail
                    f_cell = f_row[idx] if idx < len(f_row) else None
                    if f_cell is not None:
                        is_formula = (getattr(f_cell, "data_type", None) == "f") or (
                            isinstance(getattr(f_cell, "value", None), str) and str(f_cell.value).startswith("=")
                        )
                        if is_formula:
                            formula_cell_count += 1
                            if v is None:
                                missing_cached_count += 1

                if row_has_any:
                    row_count += 1

        preview = XlsxSheetPreview(
            name=name,
            row_count=row_count,
            columns=columns,
            non_empty_counts=non_empty_counts,
            formula_cell_count=formula_cell_count,
            formula_missing_cached_value_count=missing_cached_count,
            truncated=truncated,
        )
        sheets.append(preview)

        if name == primary_sheet_name:
            primary_columns = columns
            primary_row_count = row_count
            primary_non_empty = dict(non_empty_counts)
            primary_formula_count = formula_cell_count
            primary_missing_cached = missing_cached_count

    return XlsxPreviewSummary(
        format="xlsx",
        primary_sheet=primary_sheet_name,
        columns=primary_columns,
        row_count=primary_row_count,
        non_empty_counts=primary_non_empty,
        formula_cell_count=primary_formula_count,
        formula_missing_cached_value_count=primary_missing_cached,
        sheets=sheets,
    )


def import_xlsx_file(
    *,
    project_root: Path,
    file_path: Path,
    mapping_ref: Optional[Dict[str, Any]] = None,
    ingest_policy_digest: Optional[str] = None,
    normalization_defaults_digest: Optional[str] = None,
    extractor_version: str = "xlsx_decoder@0.1.0",
    brownfield_conn: Optional[sqlite3.Connection] = None,
    artifact_store: Optional[ArtifactStore] = None,
) -> str:
    """XLSX 파일을 import하고 ImportRun(IMP-*)를 반환합니다."""

    conn = brownfield_conn or open_brownfield_db(project_root=project_root)
    migrate_brownfield_db(conn)

    art_store = artifact_store or ArtifactStore(project_root=project_root)
    upload_artifact_id = art_store.put_file(Path(file_path), kind="upload", mime_type=_XLSX_MIME, dedupe=True)
    stored_path = art_store.get_path(upload_artifact_id)
    if stored_path is None:
        raise RuntimeError("Failed to resolve stored artifact path")

    preview = _decode_xlsx_to_preview(stored_path)

    imp_store = ImportRunStore(conn)
    uow = UnitOfWork(conn)

    with uow.transaction():
        imp_id = imp_store.create_staged(
            artifact_ids=[upload_artifact_id],
            mapping_ref=mapping_ref,
            extractor_version=extractor_version,
            ingest_policy_digest=ingest_policy_digest,
            normalization_defaults_digest=normalization_defaults_digest,
        )

        preview_artifact_id = art_store.put_json(
            preview.to_dict(),
            kind="brownfield_preview",
            meta={"import_run_id": imp_id, "upload_artifact_id": upload_artifact_id},
        )
        imp_store.attach_preview(imp_id, preview_artifact_id)

    return imp_id
